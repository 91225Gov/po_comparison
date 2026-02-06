"""
Retail Analytics Manager - Excel File Comparison Application
Upload two Excel files, run verification, and view column-field wise differences.
"""

import streamlit as st
import pandas as pd
from io import BytesIO
import html

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from excel_compare import compare_excel_files, ComparisonResult, CellDifference, KEY_COLUMN

st.set_page_config(
    page_title="Retail Analytics – Excel Compare",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Custom styling for a professional retail analytics look
st.markdown("""
<style>
    .main-header {
        font-size: 1.8rem;
        font-weight: 700;
        color: #1e3a5f;
        margin-bottom: 0.5rem;
    }
    .sub-header {
        color: #5a6c7d;
        font-size: 1rem;
        margin-bottom: 2rem;
    }
    .metric-card {
        background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%);
        padding: 1rem 1.25rem;
        border-radius: 8px;
        border-left: 4px solid #3b82f6;
        margin: 0.5rem 0;
    }
    .diff-row {
        padding: 0.5rem;
        border-radius: 4px;
    }
    .stButton > button {
        font-weight: 600;
        border-radius: 8px;
    }
</style>
""", unsafe_allow_html=True)


def load_excel(uploaded_file) -> tuple[dict[str, pd.DataFrame], list[str]]:
    """Load all sheets from an Excel file. Returns dict of sheet_name -> DataFrame and list of sheet names."""
    xl = pd.ExcelFile(uploaded_file)
    sheets = {}
    for name in xl.sheet_names:
        sheets[name] = pd.read_excel(xl, sheet_name=name, header=0)
    return sheets, xl.sheet_names


def run_comparison(
    file1_sheets: dict[str, pd.DataFrame],
    file2_sheets: dict[str, pd.DataFrame],
    sheet1_name: str,
    sheet2_name: str,
    key_columns: list[str],
) -> ComparisonResult:
    df1 = file1_sheets[sheet1_name]
    df2 = file2_sheets[sheet2_name]
    return compare_excel_files(df1, df2, sheet1_name, sheet2_name, key_columns=key_columns)


def _safe_str(val) -> str:
    """Convert value to string for HTML display."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip()


def _single_crosstab_to_html(key_crosstabs: list, key_column_label: str, field_columns: list) -> str:
    """
    One table: rows = keys (e.g. PO numbers), columns = key + for each field two sub-columns (File 1, File 2)
    + last column = count of columns that differ. Fields with a difference are highlighted: red background, white font.
    """
    lines = [
        '<table style="border-collapse: collapse; width: 100%; margin-bottom: 1rem;">',
        "<thead><tr style=\"background: #1e3a5f; color: white;\">",
        f'<th style="border: 1px solid #ddd; padding: 8px; text-align: left;">{html.escape(key_column_label)}</th>',
    ]
    for col in field_columns:
        col_esc = html.escape(col)
        lines.append(f'<th style="border: 1px solid #ddd; padding: 8px; text-align: center;" colspan="2">{col_esc}</th>')
    lines.append(f'<th style="border: 1px solid #ddd; padding: 8px; text-align: center;">No. of columns differing</th>')
    lines.append("</tr><tr style=\"background: #2c5282; color: white;\">")
    lines.append('<th style="border: 1px solid #ddd; padding: 6px;"></th>')
    for _ in field_columns:
        lines.append('<th style="border: 1px solid #ddd; padding: 6px; text-align: center;">File 1</th>')
        lines.append('<th style="border: 1px solid #ddd; padding: 6px; text-align: center;">File 2</th>')
    lines.append('<th style="border: 1px solid #ddd; padding: 6px;"></th>')
    lines.append("</tr></thead><tbody>")

    for tab in key_crosstabs:
        key_esc = html.escape(_safe_str(tab["key_value"]))
        by_col = {r["column"]: r for r in tab["rows"]}
        diff_count = sum(1 for r in tab["rows"] if r.get("is_difference", False))
        lines.append("<tr>")
        lines.append(f'<td style="border: 1px solid #ddd; padding: 8px; font-weight: 600;">{key_esc}</td>')
        for col in field_columns:
            r = by_col.get(col, {})
            f1 = html.escape(_safe_str(r.get("file1", "")))
            f2 = html.escape(_safe_str(r.get("file2", "")))
            is_diff = r.get("is_difference", False)
            style = 'border: 1px solid #ddd; padding: 8px; background-color: #c0392b; color: white;' if is_diff else 'border: 1px solid #ddd; padding: 8px;'
            lines.append(f'<td style="{style}">{f1}</td>')
            lines.append(f'<td style="{style}">{f2}</td>')
        lines.append(f'<td style="border: 1px solid #ddd; padding: 8px; font-weight: 600; text-align: center;">{diff_count}</td>')
        lines.append("</tr>")
    lines.append("</tbody></table>")
    return "\n".join(lines)


def _crosstab_to_dataframe(key_crosstabs: list, key_column_label: str, field_columns: list) -> pd.DataFrame:
    """Build DataFrame from crosstab (one row per key, sub-columns File 1/File 2 per field, last col = variance count)."""
    rows = []
    for tab in key_crosstabs:
        by_col = {r["column"]: r for r in tab["rows"]}
        diff_count = sum(1 for r in tab["rows"] if r.get("is_difference", False))
        row = {key_column_label: tab["key_value"]}
        for col in field_columns:
            r = by_col.get(col, {})
            row[f"{col} (File 1)"] = _safe_str(r.get("file1", ""))
            row[f"{col} (File 2)"] = _safe_str(r.get("file2", ""))
        row["No. of columns differing"] = diff_count
        rows.append(row)
    return pd.DataFrame(rows)


# Red background and white font for difference cells in Excel
_EXCEL_DIFF_FILL = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
_EXCEL_DIFF_FONT = Font(color="FFFFFF", bold=False)


def _write_crosstab_excel_with_formatting(
    key_crosstabs: list,
    key_column_label: str,
    field_columns: list,
) -> BytesIO:
    """Write crosstab to Excel and apply red background + white font to cells where File 1 != File 2."""
    crosstab_df = _crosstab_to_dataframe(key_crosstabs, key_column_label, field_columns)
    buf = BytesIO()
    crosstab_df.to_excel(buf, index=False, sheet_name="Areas of difference")
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    # Data rows: Excel row 2 = first data row (row index 0)
    for r, tab in enumerate(key_crosstabs):
        excel_row = r + 2
        by_col = {row["column"]: row for row in tab["rows"]}
        for c, col in enumerate(field_columns):
            cell_info = by_col.get(col, {})
            if cell_info.get("is_difference", False):
                # File 1 and File 2 columns for this field (1-based)
                col_file1 = 2 + c * 2
                col_file2 = 3 + c * 2
                for col_idx in (col_file1, col_file2):
                    cell = ws.cell(row=excel_row, column=col_idx)
                    cell.fill = _EXCEL_DIFF_FILL
                    cell.font = _EXCEL_DIFF_FONT
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def main():
    st.markdown('<p class="main-header">📊 Retail Analytics Manager – Excel Comparison</p>', unsafe_allow_html=True)
    st.markdown(
        '<p class="sub-header">Compare two Excel files by matching rows on the unique column(s) you specify. Upload files, set the key column(s), then run verification.</p>',
        unsafe_allow_html=True,
    )

    # Unique column(s) for matching rows — at the top so user sets it first
    key_input = st.text_input(
        "Unique column(s) for matching rows",
        value=KEY_COLUMN,
        placeholder="e.g. Purchasing document number  or  Column1, Column2, Column3",
        help="Enter one column name, or several separated by commas. Rows are matched when all these columns have the same values in both files.",
    )
    key_columns = [c.strip() for c in key_input.split(",") if c.strip()]

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("File 1")
        file1 = st.file_uploader("Choose first Excel file", type=["xlsx", "xls"], key="file1")

    with col2:
        st.subheader("File 2")
        file2 = st.file_uploader("Choose second Excel file", type=["xlsx", "xls"], key="file2")

    file1_sheets = file2_sheets = None
    sheet_names1 = sheet_names2 = []

    if file1:
        try:
            file1_sheets, sheet_names1 = load_excel(file1)
            st.caption(f"File 1: **{file1.name}** — {len(sheet_names1)} sheet(s)")
        except Exception as e:
            st.error(f"Could not read File 1: {e}")

    if file2:
        try:
            file2_sheets, sheet_names2 = load_excel(file2)
            st.caption(f"File 2: **{file2.name}** — {len(sheet_names2)} sheet(s)")
        except Exception as e:
            st.error(f"Could not read File 2: {e}")

    sheet1_name = sheet2_name = None
    if file1_sheets and sheet_names1:
        sheet1_name = st.sidebar.selectbox("Sheet in File 1", sheet_names1, key="sheet1")
    if file2_sheets and sheet_names2:
        sheet2_name = st.sidebar.selectbox("Sheet in File 2", sheet_names2, key="sheet2")

    st.sidebar.markdown("---")
    st.sidebar.caption("Key: **" + (", ".join(key_columns) if key_columns else KEY_COLUMN) + "**")
    execute = st.sidebar.button("▶ Run comparison", type="primary", use_container_width=True)

    if execute:
        if not file1 or not file2:
            st.warning("Please upload both Excel files before running the comparison.")
        elif not file1_sheets or not file2_sheets:
            st.warning("Could not load one or both files. Check file format.")
        elif not key_columns:
            st.warning("Please enter at least one unique column name for matching rows.")
        else:
            with st.spinner("Running verification..."):
                result = run_comparison(
                    file1_sheets, file2_sheets,
                    sheet1_name or sheet_names1[0],
                    sheet2_name or sheet_names2[0],
                    key_columns=key_columns,
                )

            if result.error:
                st.error(result.error)
                return

            st.success("Verification complete.")
            st.markdown("---")
            st.subheader("Summary")

            summary = result.summary_dict()
            cols = st.columns(3)
            for idx, (k, v) in enumerate(summary.items()):
                with cols[idx % 3]:
                    st.metric(k, v)

            st.markdown("---")
            st.subheader("Columns used for comparison")
            st.caption(f"Only these columns are compared for each key ({result.key_column}).")
            if result.columns_compared:
                st.code(", ".join(result.columns_compared))
            if result.requested_columns_missing_in_file1:
                st.warning(f"Requested columns not found in File 1: {', '.join(result.requested_columns_missing_in_file1)}")
            if result.requested_columns_missing_in_file2:
                st.warning(f"Requested columns not found in File 2: {', '.join(result.requested_columns_missing_in_file2)}")

            st.markdown("---")
            st.subheader("Areas of difference")

            key_label = result.key_column or "Unique key"
            st.caption(f"All **{key_label}** values with differences are listed in a **single table**: one row per key, one column per field. Each cell shows **File 1** and **File 2** values; cells where they differ are **red with white text**.")

            if result.keys_only_in_file1:
                st.markdown(f"**{key_label} only in File 1 (missing in File 2):**")
                st.code(", ".join(str(k) for k in result.keys_only_in_file1))
            if result.keys_only_in_file2:
                st.markdown(f"**{key_label} only in File 2 (not in File 1):**")
                st.code(", ".join(str(k) for k in result.keys_only_in_file2))

            if result.columns_only_in_file1:
                st.markdown("**Columns only in File 1:**")
                st.code(", ".join(result.columns_only_in_file1))
            if result.columns_only_in_file2:
                st.markdown("**Columns only in File 2:**")
                st.code(", ".join(result.columns_only_in_file2))

            if result.differences:
                # Single table: all PO numbers as rows, fields as columns; each cell = File 1 & File 2 values; red = variance
                st.markdown("**All purchase order numbers — fields as crosstab (File 1 / File 2 per column; red = variance):**")
                st.markdown(
                    _single_crosstab_to_html(
                        result.key_crosstabs,
                        key_label,
                        result.columns_compared,
                    ),
                    unsafe_allow_html=True,
                )

                buf = _write_crosstab_excel_with_formatting(
                    result.key_crosstabs,
                    key_label,
                    result.columns_compared,
                )
                st.download_button(
                    "Download Areas of difference as Excel",
                    data=buf,
                    file_name="excel_comparison_areas_of_difference.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                if result.common_columns:
                    st.info("No cell differences found. For every key in File 1, the matching row in File 2 has the same values in common columns.")
                else:
                    st.info("No common columns to compare.")

    else:
        st.info("Upload two Excel files and click **Run comparison** to see differences.")


if __name__ == "__main__":
    main()
